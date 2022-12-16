import os
import re
import openpyxl as xl

##archivo concentrador

pth2 = os.listdir("../")
print(pth2[6])
libro = xl.load_workbook(filename='../' + pth2[6])
stn2 = libro.sheetnames
predio = libro[stn2[3]]
produccion = libro[stn2[4]]
cultivos = libro[stn2[5]]
indicadores = libro[stn2[6]]
bio = libro[stn2[7]]
cosecha = libro[stn2[8]]
financiamiento = libro[stn2[9]]


##ordena las listas

def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
    return sorted(data, key=alphanum_key)


# obtener direcciones de los 34 archivos
pth = '../ALL/'
content = sorted_alphanumeric(os.listdir(pth))
print(content)
# obtener nombres de los hojas
d = xl.load_workbook(pth + content[0])
stn = d.sheetnames


##aqui obtienes un datos que quieres de las 34 y los guardas en una lista

def obtenerDatos(casilla, nsheet):
    m = []

    for file in content:
        df = xl.load_workbook(pth + file)
        st = df[stn[nsheet]]
        mo = st[casilla].value
        m.append(mo)
        df.close()
    print(m)
    return m


##ingresar los datos de la lista al archivo concentrador
def ORP(m):
    j = 249
    for i in m:

        k = str(j)
        if type(i) == str:
            pass
        elif i == None:
            pass
        elif i < -50:
            produccion['C' + k] = i
        elif -50 <= i < 0:
            produccion['D' + k] = i
        elif 0 <= i < 50:
            produccion['E' + k] = i
        elif 50 <= i < 100:
            produccion['F' + k] = i
        elif 100 <= i < 150:
            produccion['G' + k] = i
        elif 150 <= i < 200:
            produccion['H' + k] = i
        elif 200 <= i < 250:
            produccion['I' + k] = i
        elif i > 250:
            produccion['J' + k] = i
        else:
            pass
        j = j + 1

    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def MO(m):
    j = 154
    for i in m:

        k = str(j)
        if type(i) == str:
            pass
        elif i == None:
            pass
        elif i < 2:
            produccion['C' + k] = i
        elif 2 <= i < 3:
            produccion['D' + k] = i
        elif 3 <= i < 4:
            produccion['E' + k] = i
        elif 4 <= i < 5:
            produccion['F' + k] = i
        elif i > 5:
            produccion['G' + k] = i
        else:
            pass
        j = j + 1

    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def conduc(m):
    j = 202
    for i in m:

        k = str(j)
        if type(i) == str:
            pass
        elif i == None:
            pass
        elif i < 1:
            produccion['C' + k] = i
        elif 1 <= i < 2:
            produccion['D' + k] = i
        elif 2 <= i < 4:
            produccion['E' + k] = i
        elif 4 <= i < 8:
            produccion['F' + k] = i
        elif 8 <= i < 16:
            produccion['G' + k] = i

        elif i > 16:
            produccion['H' + k] = i
        else:
            pass
        j = j + 1

    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def pH(m):
    j = 296
    for i in m:

        k = str(j)
        if type(i) == str:
            pass
        elif i == None:
            pass
        elif i < 6:
            produccion['C' + k] = i
        elif 6 <= i < 6.5:
            produccion['D' + k] = i
        elif 6.6 <= i < 7.2:
            produccion['E' + k] = i
        elif 7.3 <= i < 7.9:
            produccion['F' + k] = i
        elif i > 8:
            produccion['G' + k] = i
        else:
            pass
        j = j + 1

    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def compactacion(m):
    j = 344
    for i in m:

        k = str(j)
        if type(i) == str:
            pass
        elif i == None:
            pass
        elif i < 100:
            produccion['C' + k] = i
        elif 100 <= i <= 200:
            produccion['D' + k] = i
        elif 200 < i <= 300:
            produccion['E' + k] = i

        elif i > 300:
            produccion['F' + k] = i
        else:
            pass
        j = j + 1

    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def numero_practicas():
    m = []

    practicas = [51, 52, 53, 54, 55]
    for file in content:
        n = 0
        df = xl.load_workbook(pth + file)
        st = df[stn[2]]

        for p in practicas:
            if st['C' + str(p)].value is not None:
                n = n + 1

        m.append(n)
        df.close()
    print("este es el numero de praticas: ", m)
    j = 507

    for i in m:
        k = str(j)
        if i == 0:
            produccion['C' + k] = 'X'
        elif i == 1:
            produccion['D' + k] = 'X'
        elif i == 2:
            produccion['E' + k] = 'X'
        elif i == 3:
            produccion['F' + k] = 'X'
        elif i == 4:
            produccion['G' + k] = 'X'
        elif i == 5:
            produccion['H' + k] = 'X'
        j = j + 1
    # guardar el archivo concentrador
    libro.save(filename='../' + pth2[6])
    m.clear()


def cultivo():
    m = []
    cultivo = ''
    practicas = [57, 58, 59, 60, 61]
    j = 556
    for file in content:
        n = 0
        df = xl.load_workbook(pth + file)
        st = df[stn[2]]

        for p in practicas:
            if st['C' + str(p)].value is not None:
                cultivo = st['C' + str(p)].value
                m.append(cultivo)
        df.close()
        k = str(j)
        for i in m:
            if i.lower() == 'maíz' or i.lower() == 'maiz':
                produccion['C' + k] = 'X'
            elif i.lower() == 'frijol':
                produccion['D' + k] = 'X'
            elif i.lower() == 'arroz':
                produccion['E' + k] = 'X'
            elif i.lower() == 'trigo':
                produccion['F' + k] = 'X'
            elif i.lower() == 'milpa':
                produccion['G' + k] = 'X'
            else:
                pass

        # guardar el archivo concentrador
        libro.save(filename='../' + pth2[6])
        m.clear()
        j = j + 1


def plagas():
    m = []
    plaga = ''
    practicas = [63, 64, 65, 66, 67]
    j = 603
    for file in content:
        n = 0
        df = xl.load_workbook(pth + file)
        st = df[stn[2]]

        for p in practicas:
            if st['C' + str(p)].value is not None:
                plaga = st['C' + str(p)].value
                m.append(plaga)
        df.close()
        k = str(j)
        if len(m) == 0:
            produccion['C' + k] = 'X'
        else:
            for i in m:
                if i.lower() == 'gusano cogollero' or i.lower() == 'gusano cogoyero':
                    produccion['D' + k] = 'X'
                elif i.lower() == 'gusano medidor':
                    produccion['E' + k] = 'X'
                elif i.lower() == 'araña roja':
                    produccion['F' + k] = 'X'
                elif i.lower() == 'pajaros' or i.lower() == 'pajáros':
                    produccion['G' + k] = 'X'
                else:
                    produccion['H' + k] = 'X'

        # guardar el archivo concentrador
        libro.save(filename='../' + pth2[6])
        m.clear()
        j = j + 1


def enfermedades():
    m = []
    plaga = ''
    practicas = [68, 69, 70, 71, 72]
    j = 650
    for file in content:
        n = 0
        df = xl.load_workbook(pth + file)
        st = df[stn[2]]

        for p in practicas:
            if st['C' + str(p)].value is not None:
                plaga = st['C' + str(p)].value
                m.append(plaga)
        df.close()
        k = str(j)
        if len(m) == 0:
            produccion['C' + k] = 'X'
        else:
            for i in m:
                if i.lower() == 'nematodos' or i.lower() == 'nemátodos':
                    produccion['D' + k] = 'X'
                elif i.lower() == 'fusarium':
                    produccion['E' + k] = 'X'
                elif i.lower() == 'mancha de asfalto':
                    produccion['F' + k] = 'X'
                elif i.lower() == 'phytopthora':
                    produccion['G' + k] = 'X'
                else:
                    produccion['H' + k] = 'X'

        # guardar el archivo concentrador
        libro.save(filename='../' + pth2[6])
        m.clear()
        j = j + 1

def fenomenos():
    m = []
    plaga = ''
    practicas = [73, 74, 75, 76, 77]
    j = 699
    for file in content:
        n = 0
        df = xl.load_workbook(pth + file)
        st = df[stn[2]]

        for p in practicas:
            if st['C' + str(p)].value is not None:
                plaga = st['C' + str(p)].value
                m.append(plaga)
        df.close()
        k = str(j)
        if len(m) == 0:
            produccion['C' + k] = 'X'
        else:
            for i in m:
                if i.lower() == 'sequia' or i.lower() == 'sequía':
                    produccion['D' + k] = 'X'
                elif i.lower() == 'heladas':
                    produccion['E' + k] = 'X'
                elif i.lower() == 'granizo':
                    produccion['F' + k] = 'X'
                elif i.lower() == 'excesos de humedad' or i.lower() == 'excesos de humedad':
                    produccion['G' + k] = 'X'
                else:
                    produccion['H' + k] = 'X'

        # guardar el archivo concentrador
        libro.save(filename='../' + pth2[6])
        m.clear()
        j = j + 1
