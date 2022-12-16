import openpyxl as xl
import os

##AQUI CARGAMOS EL ARCHIVO DONDE ESTAN LA CURPS
pth2 = os.listdir("../")
print(pth2[6])

libro = xl.load_workbook(filename='../' + pth2[6])
stn2 = libro.sheetnames
lcurps = libro[stn2[2]]
curps = []
j = 12
for i in range(34):
    k = str(j)
    c = lcurps['F' + k].value
    curps.append(c)
    j += 1

print(curps[33])

# obtener direcciones de los 34 archivos
pth = '../IN/'
content = os.listdir(pth)
print(content)
# obtener nombres de los hojas
d = xl.load_workbook(pth + content[0])
stn = d.sheetnames
print(stn)

for element in content:
    archivo = xl.load_workbook(pth + element)
    st = archivo[stn[0]]
    cp = st['B14'].value
    i = 0
    for c in curps:
        if cp == c:
            k = str(i+1)
            archivo.save(filename='../OUT/' + k + '.xlsx')
        i += 1
    archivo.close()