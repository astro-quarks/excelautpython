import os
import re
import openpyxl as xl

##archivo concentrador

pth2 = os.listdir('../')
libro = xl.load_workbook(filename='../' + pth2[5])



stn2 = libro.sheetnames


##ordena las listas

def sorted_alphanumeric(data):
    convert = lambda text: int(text) if text.isdigit() else text.lower()
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
    return sorted(data, key=alphanum_key)


# obtener direcciones de los 34 archivos
pth = '../ALL/'
content = sorted_alphanumeric(os.listdir(pth))
print(content)
# obtener nombres de los hojas
d = xl.load_workbook(pth + content[0])
stn = d.sheetnames

##aqui obtienes un datos que quieres de las 34 y los guardas en una lista
m = []
for file in content:
    df = xl.load_workbook(pth + file)
    st = df[stn[2]]
    mo = st['C41'].value
    m.append(mo)
    df.close()

print(m)

##ingresar los datos de la lista al archivo concentrador
j = 249
for i in m:

    k = str(j)
    if type(i) == str:
        pass
    elif i == None:
        pass
    elif i < -50:
        produccion['C' + k] = i
    elif -50 <= i < 0:
        produccion['D' + k] = i
    elif 0 <= i < 50:
        produccion['E' + k] = i
    elif 50 <= i < 100:
        produccion['F' + k] = i
    elif 100 <= i < 150:
        produccion['G' + k] = i
    elif 150 <= i < 200:
        produccion['H' + k] = i
    elif 200 <= i < 250:
        produccion['I' + k] = i
    elif i > 250:
        produccion['J' + k] = i
    else:
        pass
    j = j + 1

# guardar el archivo concentrador
libro.save(filename='../' + pth2[5])
